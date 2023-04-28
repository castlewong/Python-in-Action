
#sheet name wrong
import openpyxl
import docx

# Open the docx file
doc = docx.Document('dict.docx')

# Create a new workbook
wb = openpyxl.Workbook()

# Loop through each table in the docx file and add it to a new sheet in the workbook
for i, table in enumerate(doc.tables):
    # Create a new sheet for each table
    sheet_name = f'Table {i+1}'
    ws = wb.create_sheet(sheet_name)

    # Loop through each row in the table and add it to the sheet
    for j, row in enumerate(table.rows):
        # Assume first column is key and second column is value.
        key = row.cells[0].text.strip()
        value = row.cells[1].text.strip()

        # Write key-value pairs to the sheet
        ws.cell(row=j+1, column=1, value=key)
        ws.cell(row=j+1, column=2, value=value)

# Save the workbook
wb.save('dict_items.xlsx')
