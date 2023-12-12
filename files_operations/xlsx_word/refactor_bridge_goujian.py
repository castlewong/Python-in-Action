# import openpyxl
#
# # Open the existing Excel file
# excel_file_path = 'csh3.xlsx'
# workbook = openpyxl.load_workbook(excel_file_path)
#
# # Select the active sheet
# sheet = workbook.active
#
# # Create new columns (columns 3 and 4) as copies of columns 1 and 2
# for row in range(1, sheet.max_row + 1):
#     sheet.cell(row=row, column=3, value=sheet.cell(row=row, column=1).value)
#     sheet.cell(row=row, column=4, value=sheet.cell(row=row, column=2).value)
#
# # Insert a new row at the top as the header
# header_row = ['table1', 'table2', 'table3', 'table4']
# sheet.insert_rows(1)
# for col, header in enumerate(header_row, start=1):
#     sheet.cell(row=1, column=col, value=header)
#
# # Save the modified Excel file
# workbook.save('csh3output.xlsx')

import os
import openpyxl

# Specify the folder containing Excel files
folder_path = '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/masBRidge'
output_path = '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/masBRidge2'
# Loop through all files in the folder
for filename in os.listdir(folder_path):
    # Check if the file is an Excel file
    if filename.endswith('.xlsx'):
        # Construct the full path to the Excel file
        excel_file_path = os.path.join(folder_path, filename)

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)

        # Select the active sheet
        sheet = workbook.active

        # Create new columns (columns 3 and 4) as copies of columns 1 and 2
        for row in range(1, sheet.max_row + 1):
            sheet.cell(row=row, column=3, value=sheet.cell(row=row, column=1).value)
            sheet.cell(row=row, column=4, value=sheet.cell(row=row, column=2).value)

        # Insert a new row at the top as the header
        header_row = ['构件编号', '构件名称', '业主系统编号', '构件位置']
        sheet.insert_rows(1)
        for col, header in enumerate(header_row, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Save the modified Excel file
        output_file_path = os.path.join(output_path, f'{os.path.splitext(filename)[0]}桥梁构件导入模板.xlsx')
        workbook.save(output_file_path)

