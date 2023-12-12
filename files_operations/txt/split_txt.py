import openpyxl

# Read the text file
text_file_path = '供水大用户.txt'
with open(text_file_path, 'r', encoding='utf-8') as file:
    lines = file.readlines()

# Create an Excel workbook and add data to it
workbook = openpyxl.Workbook()
sheet = workbook.active

# Write lines to Excel sheet
for row, line in enumerate(lines, start=1):
    parts = [part.strip() for part in line.split('；') + line.split('、')]
    for col, part in enumerate(parts, start=1):
        # sheet.cell(row=row, column=col, value=part)
        sheet.cell(row=row + col - 1, column=1, value=part)


# Save the Excel file
excel_file_path = '供水大用户output.xlsx'
workbook.save(excel_file_path)

