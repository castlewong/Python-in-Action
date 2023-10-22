import openpyxl
import shutil
import difflib
import pandas as pd

def compare_arrays(arrayDB, arrayProvince):
    result = []
    for prov_element in arrayProvince:
        if prov_element in arrayDB:
            result.append("not-in-DB-type")
        else:
            found_similar = False
            for db_element in arrayDB:
                similarity_ratio = difflib.SequenceMatcher(None, prov_element, db_element).ratio()
                if similarity_ratio > 0.6:  # Adjust this threshold as needed
                    result.append("similar-to-DBtype")
                    found_similar = True
                    break
            if not found_similar:
                result.append("not-in-DB-type")
    print(result)
    return result

# 获取所有的省平台字典类型
def get_sheet_names(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the names of all sheets in the workbook
    sheet_names_fromProvince = workbook.sheetnames

    # print(sheet_names_fromProvince)

    return sheet_names_fromProvince

# 获取所有库里的字典类型
def get_xlsx_files(directory):
    # Get a list of all the files in the directory
    files = os.listdir(directory)

    # Filter the list to only include files with the .xlsx extension
    xlsx_files = [file for file in files if file.endswith('.xlsx')]

    # Remove the extension from the file names
    xlsx_files_without_extension_fromDB = [os.path.splitext(file)[0] for file in xlsx_files]

    # print(xlsx_files_without_extension_fromDB)

    return xlsx_files_without_extension_fromDB

def rename_sheets_xlsx(filepath, new_name):
    wb = openpyxl.load_workbook(filepath)
    for sheetname in wb.sheetnames:
        if '字典' in sheetname:
            new_sheetname = sheetname.replace('字典', '')
            wb[sheetname].title = new_sheetname
    wb.save(new_name)

# split xlsx to xlsxs by sheets

# rename_sheets_xlsx('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/dict_items_province.xlsx', 'dict_items_province_new.xlsx')
# arrPro = get_sheet_names(
#     '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/dict_items_group.xlsx')
# arrDB = get_xlsx_files('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/xlsx-to-combline')
#
# compare_arrays(arrPro, arrDB)

def save_sheet_names(filename):
    # Load the workbook using openpyxl
    workbook = openpyxl.load_workbook(filename)

    # Get a list of all sheet names
    sheet_names = workbook.sheetnames

    # Create a new workbook
    new_workbook = openpyxl.Workbook()

    # Get the active worksheet
    active_sheet = new_workbook.active

    # Set the first cell to "Sheet Names"
    active_sheet['A1'] = "Sheet Names"

    # Write the sheet names to the worksheet
    for i, sheet_name in enumerate(sheet_names):
        active_sheet.cell(row=i + 2, column=1, value=sheet_name)

    # Save the workbook
    new_workbook.save('sheet_names.xlsx')

# save_sheet_names('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/combined.xlsx')


def group_data_to_excel_single_file(filename, group_col):
    """
    Groups a pandas DataFrame by a specified column and exports each group's data to a separate Excel file.

    Args:
        df (pandas.DataFrame): The DataFrame to group and export.
        group_col (str): The name of the column to group the DataFrame by.

    Returns:
        None
    """
    df = pd.read_excel(filename)

    # Group the data by the specified column
    grouped_data = df.groupby(group_col)

    # Iterate over each group and write its data to a new Excel file
    for group_name, group_df in grouped_data:
        # Create a new Excel file with a sheet for this group
        writer = pd.ExcelWriter(f'{group_name}.xlsx', engine='xlsxwriter')

        # Write the group's data to the new sheet
        group_df.to_excel(writer, index=False)

        # Close the writer to save the Excel file
        writer.close()



def transpose_all_xlsx_files(input_dir, output_dir):
    """
    Transposes (switches) the rows and columns of all sheets in multiple xlsx files.

    Args:
    input_dir (str): The directory containing the input xlsx files.
    output_dir (str): The directory where the transposed xlsx files will be saved.

    Returns:
    None
    """
    # Get a list of all the xlsx files in the input directory
    xlsx_files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx')]

    # Loop through each xlsx file
    for file in xlsx_files:
        # Load the xlsx file into a Pandas DataFrame
        df = pd.read_excel(os.path.join(input_dir, file), sheet_name=None)

        # Transpose all sheets in the DataFrame
        transposed_df = {sheet_name: df[sheet_name].T for sheet_name in df.keys()}

        # Save the transposed xlsx file to the output directory
        writer = pd.ExcelWriter(os.path.join(output_dir, file), engine='xlsxwriter')
        for sheet_name, sheet_df in transposed_df.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
        writer.save()

from openpyxl import load_workbook
# def modify_xlsx_files(folder_path):
#     # Loop through all xlsx files in the folder
#     for file_name in os.listdir(folder_path):
#         if file_name.endswith('.xlsx'):
#             # Load the workbook
#             file_path = os.path.join(folder_path, file_name)
#             wb = load_workbook(filename=file_path)
#
#             # Delete all rows except the fourth one
#             ws = wb.active
#             for i in range(2, ws.max_row + 1):
#                 if i != 4:
#                     ws.delete_rows(i)
#             # Save the modified workbook
#             wb.save(file_path)
from openpyxl import load_workbook

def modify_xlsx_files(folder_path, output_folder_path):
    # Loop through all xlsx files in the folder
    for file_name in os.listdir(folder_path):
        if file_name.endswith('.xlsx'):
            # Load the workbook
            file_path = os.path.join(folder_path, file_name)
            df = pd.read_excel(file_path)

            # Find rows containing the string "主键"
            mask = df.apply(lambda row: '主键' in str(row.values), axis=1)

            # Filter the dataframe to only include rows containing "主键"
            df = df[mask]

            # Save the updated dataframe to a new Excel file in the specified directory
            basename, extension = os.path.splitext(file_name)
            new_filepath = os.path.join(output_folder_path, basename + '_modified' + extension)
            writer = pd.ExcelWriter(new_filepath, engine='xlsxwriter')
            df.to_excel(writer, index=False)
            writer.save()


import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_columns(folder_path):
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF99', fill_type='solid')
    keywords = ['地址', '名称', '经度', '纬度']

    for filename in os.listdir(folder_path):
        if filename.endswith('.xlsx'):
            file_path = os.path.join(folder_path, filename)
            wb = load_workbook(file_path)

            for sheet in wb.worksheets:
                for col in sheet.columns:
                    for cell in col:
                        if any(keyword in str(cell.value) for keyword in keywords):
                            cell.fill = yellow_fill

            wb.save(file_path)
def rename_files_in_dir(directory):
    for filename in os.listdir(directory):
        if filename.endswith('省平台') :
            new_name = filename.replace('.xlsx省平台', '省平台.xlsx')
            # Rename the file
            os.rename(os.path.join(directory, filename), os.path.join(directory, new_name))
def save_xlsx_sheets_as_files(file_path, output_folder):
    # Load the Excel file into a Pandas dataframe
    xl_file = pd.ExcelFile(file_path)

    # Create the output folder if it doesn't exist
    os.makedirs(output_folder, exist_ok=True)

    # Loop through each sheet in the dataframe and save it to a separate Excel file
    for sheet_name in xl_file.sheet_names:
        sheet_df = xl_file.parse(sheet_name)
        sheet_file_path = os.path.join(output_folder, f"{sheet_name}.xlsx")
        sheet_df.to_excel(sheet_file_path, index=False)
def find_similar_files(a_folder, b_folder, c_folder):
    # Get the list of files in folder a
    a_files = os.listdir(a_folder)

    # Get the list of files in folder b
    b_files = os.listdir(b_folder)

    # Loop through the files in folder a
    for a_file in a_files:
        # Get the first four characters of the file name
        a_prefix = a_file[:4]

        # Loop through the files in folder b
        for b_file in b_files:
            # Get the first four characters of the file name
            b_prefix = b_file[:4]

            # If the prefixes match, copy the file from folder b to folder c
            if a_prefix == b_prefix:
                src_path = os.path.join(b_folder, b_file)
                dst_path = os.path.join(c_folder, b_file)
                shutil.copyfile(src_path, dst_path)

# find_similar_files('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_fromProvince', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_fromDB', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_compareTwo')

def update_first_row_name(folder_path, name):
    for filename in os.listdir(folder_path):
        if not '省平台' in filename and filename.endswith('.xlsx'):
            filepath = os.path.join(folder_path, filename)
            df = pd.read_excel(filepath, header=None)
            if df.iloc[0][0] == name:
                df.iloc[0][0] = '名称'
                df.to_excel(filepath, index=False, header=False)


# def compare_files_in_folder(folder_path):
#     """
#     Compare xlsx files in a folder based on the '名称' column.
#     Files with similar names are compared in pairs.
#     Differences in '名称' column are highlighted in yellow.
#
#     :param folder_path: str, path of the folder containing xlsx files
#     """
#
#     # get a list of all xlsx files in the folder
#     file_list = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
#
#     # create a dictionary to store dataframes for each file
#     file_dict = {}
#
#     # loop through each file and read it into a dataframe
#     for filename in file_list:
#         file_path = os.path.join(folder_path, filename)
#         df = pd.read_excel(file_path)
#         # only keep the first four characters of the filename as the dictionary key
#         key = filename[:4]
#         file_dict[key] = df
#
#     # loop through each pair of dataframes with similar keys and compare their '名称' columns
#     for key1, df1 in file_dict.items():
#         for key2, df2 in file_dict.items():
#             if key1 == key2:
#                 continue
#             elif key1 == key2[:4]:
#                 # compare the '名称' columns and highlight differences in yellow
#                 diff_df = pd.concat([df1['名称'], df2['名称']], axis=1, keys=['file1', 'file2'])
#                 diff_df['highlight'] = ''
#                 mask = (diff_df['file1'] != diff_df['file2'])
#                 diff_df.loc[mask, 'highlight'] = 'background-color: yellow'
#                 display(diff_df.style.apply(lambda x: x.highlight_max(axis=0), axis=0).set_properties(
#                     **{'text-align': 'left'}).set_table_styles(
#                     [{'selector': '.row_heading', 'props': [('display', 'none')]}]).apply(
#                     lambda x: diff_df['highlight'], axis=1))

# import pandas as pd
# import os
#
# def compare_similar_xlsx_files(directory):
#     # Get a list of all the xlsx files in the directory
#     files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
#
#     # Create a dictionary to store the dataframes for each file
#     dfs = {}
#     for file in files:
#         # Extract the first four characters from the filename to use as a key
#         key = file[:4]
#         # Read the file into a dataframe and store it in the dictionary
#         dfs.setdefault(key, pd.read_excel(os.path.join(directory, file)))
#
#     # Loop through the keys in the dictionary, comparing dataframes with matching keys
#     for key in dfs:
#         # Get a list of keys that are similar to the current key
#         similar_keys = [k for k in dfs.keys() if k != key and k[:4] == key[:4]]
#         for similar_key in similar_keys:
#             # Compare the '名称' column in each dataframe
#             diff = dfs[key]['名称'] != dfs[similar_key]['名称']
#             # Highlight any cells that are different in yellow
#             # Define the red fill style
#             red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
#
#             # Highlight any cells that are different in red
#             for cell in dfs[key].loc[diff, '名称'].index:
#                 ws = wb[key]
#                 ws[f'名称{cell + 1}'].fill = red_fill
#
#             for cell in dfs[similar_key].loc[diff, '名称'].index:
#                 ws = wb[similar_key]
#                 ws[f'名称{cell + 1}'].fill = red_fill
#
#     # Write the modified dataframes back to their respective files
#     for key, df in dfs.items():
#         df.to_excel(os.path.join(directory, f'{key}_output.xlsx'), index=False)


# update_first_row_name('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_compareTwo', 'name')

# save_xlsx_sheets_as_files('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_fromProvince.xlsx', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_fromProvince')

# rename_files_in_dir('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_fromProvince')

# highlight_columns('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/公共导出2')

# modify_xlsx_files('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/公共导出', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/公共导出2')

# transpose_all_xlsx_files('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/公共数据表格式', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/公共导出')

# group_data_to_excel_single_file('public_comment.xlsx', '名称')

#
# def compare_names2(a_file_path, b_file_path):
#     # Read in both Excel files
#     a_df = pd.read_excel(a_file_path)
#     b_df = pd.read_excel(b_file_path)
#
#     # Convert the '名称' columns to sets for comparison
#     a_names = set(a_df['名称'])
#     b_names = set(b_df['名称'])
#
#     # Determine which names are in file A but not file B
#     missing_names = a_names - b_names
#
#     # Create a boolean mask for highlighting the cells with missing names in file A
#     mask = a_df['名称'].apply(lambda x: x in missing_names)
#
#     # Output a copy of the DataFrame with highlighted cells
#     a_highlighted = a_df.copy()
#     a_highlighted.loc[mask, '名称'] = '<span style="background-color: yellow">' + a_highlighted.loc[
#         mask, '名称'] + '</span>'
#     return a_highlighted
#
# def compare_names(a_file_path, b_file_path, output_file_path):
#     # Read in both Excel files
#     a_df = pd.read_excel(a_file_path)
#     b_df = pd.read_excel(b_file_path)
#
#     # Convert the '名称' columns to sets for comparison
#     a_names = set(a_df['名称'])
#     b_names = set(b_df['名称'])
#
#     # Determine which names are in file A but not file B
#     missing_names = a_names - b_names
#
#     print(missing_names)
#
#     # Create a boolean mask for highlighting the cells with missing names in file A
#     mask = a_df['名称'].apply(lambda x: x in missing_names)
#
#     # Output a copy of the DataFrame with highlighted cells
#     a_highlighted = a_df.copy()
#     a_highlighted.loc[mask, '名称'] = '<span style="background-color: yellow">' + a_highlighted.loc[
#         mask, '名称'] + '</span>'
#
#     # Save the highlighted DataFrame to a new Excel file
#     # a_highlighted.to_excel(output_file_path, index=False)

#
#
# import openpyxl
#

def compare_columns(file_a, file_b):
    # Load the Excel files
    workbook_a = openpyxl.load_workbook(filename=file_a)
    workbook_b = openpyxl.load_workbook(filename=file_b)

    # Get the sheet names
    sheet_names_a = workbook_a.sheetnames
    sheet_names_b = workbook_b.sheetnames

    # Compare the columns on each sheet
    for sheet_name in sheet_names_a:
        sheet_a = workbook_a[sheet_name]
        sheet_b = workbook_b[sheet_name]

        # Get the column index
        column_index_a = None
        for cell in sheet_a[1]:
            if cell.value == "name":
                column_index_a = cell.column
                break
        column_index_b = None
        for cell in sheet_b[1]:
            if cell.value == "name":
                column_index_b = cell.column
                break

        # Get the column values as sets
        column_a = set()
        for row in sheet_a.iter_rows(min_row=2, min_col=column_index_a, max_col=column_index_a, values_only=True):
            column_a.add(row[0])
        column_b = set()
        for row in sheet_b.iter_rows(min_row=2, min_col=column_index_b, max_col=column_index_b, values_only=True):
            column_b.add(row[0])

        # Identify the differences and create a new workbook to store them
        diff = column_a - column_b
        if len(diff) > 0:
            workbook_diff = openpyxl.Workbook()
            sheet_diff = workbook_diff.active

            # Write the differences to the new workbook
            row_num = 1
            for value in diff:
                sheet_diff.cell(row=row_num, column=1, value=value)
                row_num += 1

            # Save the new workbook
            filename_diff = f"{sheet_name}_diff.xlsx"
            workbook_diff.save(filename_diff)c

    # Close the workbooks
    workbook_a.close()
    workbook_b.close()

# compare_columns('/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_compareTwo copy2/字典相同/关联目标类型省平台.xlsx', '/Users/wilburwong/Documents/Python-in-Action/files_operations/xlsx_word/items_compareTwo copy2/字典相同/关联目标类型.xlsx')