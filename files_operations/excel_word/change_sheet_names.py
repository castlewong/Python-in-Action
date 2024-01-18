from openpyxl import load_workbook

# Load the dict_itemsBack.xlsx workbook
wb = load_workbook('dict_itemsBack.xlsx')

# Load the sheet names from dict_types.xlsx
types_wb = load_workbook('dict_type.xlsx')
types_sheet = types_wb.active
sheet_names = [cell.value for cell in types_sheet['A'] if cell.value]

# Rename each sheet in the workbook based on the names in dict_types.xlsx
for i, sheet_name in enumerate(sheet_names):
    wb.worksheets[i].title = sheet_name

# Save the updated workbook
wb.save('dict_itemsBack_updated.xlsx')
